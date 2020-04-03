'''
Васю назначили завхозом в туристической группе и он подошёл к подготовке ответственно, составив 
справочник продуктов с указанием калорийности на 100 грамм, а также содержание белков, жиров и 
углеводов на 100 грамм продукта. Ему не удалось найти всю информацию, поэтому некоторые ячейки 
остались незаполненными (можно считать их значение равным нулю). Также он использовал какой-то 
странный офисный пакет и разделял целую и дробную часть чисел запятой. Таблица доступа по ссылке 
https://stepik.org/media/attachments/lesson/245290/trekking1.xlsx
Вася хочет минимизировать вес продуктов и для этого брать самые калорийные продукты. Помогите ему 
и упорядочите продукты по убыванию калорийности. В случае, если продукты имеют одинаковую 
калорийность - упорядочите их по названию. В качестве ответа необходимо сдать названия продуктов, 
по одному в строке.
'''

import xlrd
import openpyxl

kkal = []

rb = xlrd.open_workbook('A:\\trekking1.xlsx')

# получаю два списка первый по строкам второй по столбцам
sheet = rb.sheet_by_index(0)
row_vals = [sheet.row_values(rownum) for rownum in range(sheet.nrows)]
col_vals = [sheet.col_values(colnum) for colnum in range(sheet.ncols)]

for row in row_vals:
    kkal.append(row[0:2])
del kkal[0]
kkal = sorted(kkal, key=lambda item: (-item[1], item[0]))


wb = openpyxl.Workbook()
wb.sheetnames
sheet = wb['Sheet']
i = 1
for row in kkal:
    value = row[0]
    cell = sheet.cell(row=i, column=1)
    i += 1
    cell.value = value

wb.save('A:\\otvet.xlsx')
