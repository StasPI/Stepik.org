'''
Главный бухгалтер компании "Рога и копыта" случайно удалил ведомость с начисленной зарплатой. К счастью, у 
него сохранились расчётные листки всех сотрудников. Помогите по этим расчётным листкам восстановить 
зарплатную ведомость. Архив с расчётными листками доступен по ссылке 
https://stepik.org/media/attachments/lesson/245299/rogaikopyta.zip (вы можете скачать и распаковать его 
вручную или самостоятельно научиться делать это с помощью скрипта на Питоне).
Ведомость должна содержать 1000 строк, в каждой строке должно быть указано ФИО сотрудника и, через пробел, 
его зарплата. Сотрудники должны быть упорядочены по алфавиту.
'''

import openpyxl

from requests import get
from zipfile import ZipFile
from os import listdir
from operator import itemgetter


# Создаю пути, иму имя файла.
url = 'https://stepik.org/media/attachments/lesson/245299/rogaikopyta.zip'
file_start_dir = 'A:\\task\\'
file_name = url[url.rindex('/') + 1:]
file_dir = file_start_dir + file_name
# Записываю скаченный файл на диск.
with open(file_dir, 'wb') as inf:
    inf.write(get(url).content)
# Извлекаю содержимое в нужную директорию.  
with ZipFile(file_dir, 'r') as inf:
    inf.extractall(path=file_start_dir)
# Ищу все распакованные файлы и создаю прямые ссылки на них.  
file_list = listdir(file_start_dir)
start_link = []
# Проверяю файлы на то что бы они имели верное разрешение.  
for link in file_list:
    if link[-5:].lower() == '.xlsx':
        start_link.append(file_start_dir + link)
# Читаю данные из файлов и сохраняю из них фио и сумму.
full_list = []
up_list = []
for file_link in start_link:
    wb = openpyxl.load_workbook(file_link)
    sheet_ranges = wb['Sheet']
    up_list.append(str(sheet_ranges['B2'].value))
    up_list.append(str(sheet_ranges['D2'].value))
    full_list.append(up_list)
    up_list = []
# Сортирую по первому значению то есть по фио.
full_list.sort(key=itemgetter(0))
# Создаю и записываю в файл.
wb = openpyxl.Workbook()
ws = wb.active
row_index = 0
for row in full_list:
    row_index += 1
    cell = ws.cell(row=row_index, column=1)
    cell.value = ' '.join(row)
wb.save('A:\\task\\_full.xlsx')
